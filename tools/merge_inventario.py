#!/usr/bin/env python3
"""Merge inventory counts into an Xetux bulk-load workbook.

The destination workbook is treated as the system template. The script writes
only the Cantidad column in the Inventario sheet, matching rows by exact Codigo.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import shutil
import sys
import tempfile
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


WORKER_URL = "https://operaciones-api.pablo-aranda.workers.dev"
SHEET_NAME = "Inventario"
CODE_HEADERS = {"codigo", "código"}
QTY_HEADERS = {"cantidad"}
EMPTY_COUNTS = (None, "", 0)


class MergeError(Exception):
    pass


@dataclass
class HeaderMap:
    row: int
    code_col: int
    qty_col: int


def norm_header(value: Any) -> str:
    return str(value or "").strip().lower()


def norm_code(value: Any) -> str:
    return str(value or "").strip()


def file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def find_headers(ws) -> HeaderMap:
    for row_idx in range(1, min(ws.max_row, 10) + 1):
      code_col = qty_col = None
      for col_idx in range(1, ws.max_column + 1):
          label = norm_header(ws.cell(row=row_idx, column=col_idx).value)
          if label in CODE_HEADERS:
              code_col = col_idx
          if label in QTY_HEADERS:
              qty_col = col_idx
      if code_col and qty_col:
          return HeaderMap(row=row_idx, code_col=code_col, qty_col=qty_col)
    raise MergeError(f'No encontre headers "Codigo" y "Cantidad" en hoja {ws.title}')


def coerce_count(value: Any) -> float | None:
    if value in EMPTY_COUNTS:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "":
            return None
        value = stripped.replace(",", "")
    try:
        num = float(value)
    except (TypeError, ValueError):
        raise ValueError(f"conteo no numerico: {value!r}") from None
    if num == 0:
        return None
    return num


def parse_source_xlsx(src_path: Path, sheet_name: str = SHEET_NAME) -> dict[str, Any]:
    wb = openpyxl.load_workbook(src_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise MergeError(f"No existe hoja {sheet_name!r} en fuente")
    ws = wb[sheet_name]
    headers = find_headers(ws)
    counts: dict[str, float] = {}
    duplicates: dict[str, list[float]] = {}
    invalid: dict[str, str] = {}
    skipped_empty = 0

    for row_idx in range(headers.row + 1, ws.max_row + 1):
        code = norm_code(ws.cell(row=row_idx, column=headers.code_col).value)
        if not code:
            continue
        raw_qty = ws.cell(row=row_idx, column=headers.qty_col).value
        try:
            qty = coerce_count(raw_qty)
        except ValueError as exc:
            invalid[code] = str(exc)
            continue
        if qty is None:
            skipped_empty += 1
            continue
        if code in counts:
            duplicates.setdefault(code, [counts[code]]).append(qty)
        counts[code] = qty

    return {
        "counts": counts,
        "source_headers": headers.__dict__,
        "duplicate_source_codes": duplicates,
        "invalid_source_quantities": invalid,
        "skipped_empty_counts": skipped_empty,
    }


def fetch_json(path: str, params: dict[str, str]) -> dict[str, Any]:
    url = WORKER_URL + path + "?" + urllib.parse.urlencode(params)
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "inventario-merge/1.0",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        return json.loads(response.read().decode("utf-8"))


def factor_default(template: dict[str, Any], code: str) -> float:
    explicit = (template.get("presMap") or {}).get(code) or []
    if explicit:
        return float(explicit[0].get("factor") or 1)
    unit = (template.get("unitMap") or {}).get(code) or ""
    defaults = (template.get("defaultPres") or {}).get(unit) or []
    if defaults:
        return float(defaults[0].get("factor") or 1)
    return 1.0


def parse_d1_session(almacen: str, fecha: str) -> dict[str, Any]:
    session = fetch_json("/inv/sesion", {"almacen": almacen, "fecha": fecha})
    if not session.get("ok") or not session.get("found"):
        raise MergeError(f"Sesion no encontrada: {almacen} {fecha}")
    template = fetch_json("/inv/plantilla", {"almacen": almacen})
    if not template.get("ok") or not template.get("found"):
        raise MergeError(f"Plantilla no encontrada: {almacen}")

    counts: dict[str, float] = {}
    for zone_id, zone_counts in (session.get("countsByZone") or {}).items():
        for code, qty in (zone_counts or {}).items():
            factor = (
                ((session.get("presChoiceByZone") or {}).get(zone_id) or {}).get(code)
                or factor_default(template, code)
            )
            counts[code] = counts.get(code, 0.0) + float(qty) * float(factor)

    admin_corrections = ((session.get("correctionsByZone") or {}).get("_admin") or {})
    for code, correction in admin_corrections.items():
        if correction and correction.get("qty") is not None:
            counts[code] = float(correction["qty"])

    return {
        "counts": counts,
        "session": {
            "almacen": almacen,
            "fecha": fecha,
            "templateHash": session.get("templateHash") or "",
            "manuales_count": len([m for m in (session.get("manuales") or []) if m.get("type") != "comment"]),
            "corrections_count": sum(len(z or {}) for z in (session.get("correctionsByZone") or {}).values()),
        },
        "template": {
            "templateHash": template.get("templateHash") or "",
            "updatedAt": template.get("updatedAt") or "",
        },
    }


def scan_destination(dst_path: Path, sheet_name: str = SHEET_NAME) -> dict[str, Any]:
    wb = openpyxl.load_workbook(dst_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise MergeError(f"No existe hoja {sheet_name!r} en destino")
    ws = wb[sheet_name]
    headers = find_headers(ws)
    rows_by_code: dict[str, int] = {}
    duplicates: dict[str, list[int]] = {}

    for row_idx in range(headers.row + 1, ws.max_row + 1):
        code = norm_code(ws.cell(row=row_idx, column=headers.code_col).value)
        if not code:
            continue
        if code in rows_by_code:
            duplicates.setdefault(code, [rows_by_code[code]]).append(row_idx)
        else:
            rows_by_code[code] = row_idx

    return {
        "headers": headers,
        "rows_by_code": rows_by_code,
        "duplicate_destination_codes": duplicates,
        "sheet_names": wb.sheetnames,
    }


def clone_cell_value(value: Any) -> Any:
    return value


def workbook_value_snapshot(path: Path) -> dict[str, dict[tuple[int, int], Any]]:
    wb = openpyxl.load_workbook(path, data_only=False)
    snap: dict[str, dict[tuple[int, int], Any]] = {}
    for ws in wb.worksheets:
        cells: dict[tuple[int, int], Any] = {}
        for row in ws.iter_rows():
            for cell in row:
                cells[(cell.row, cell.column)] = clone_cell_value(cell.value)
        snap[ws.title] = cells
    return snap


def diff_snapshots(
    before: dict[str, dict[tuple[int, int], Any]],
    after: dict[str, dict[tuple[int, int], Any]],
    allowed: set[tuple[str, int, int]],
) -> list[dict[str, Any]]:
    diffs = []
    for sheet in sorted(set(before) | set(after)):
        before_cells = before.get(sheet, {})
        after_cells = after.get(sheet, {})
        for coord in sorted(set(before_cells) | set(after_cells)):
            if (sheet, coord[0], coord[1]) in allowed:
                continue
            if before_cells.get(coord) != after_cells.get(coord):
                diffs.append({
                    "sheet": sheet,
                    "row": coord[0],
                    "column": coord[1],
                    "before": before_cells.get(coord),
                    "after": after_cells.get(coord),
                })
    return diffs


def apply_counts(
    dst_path: Path,
    counts: dict[str, float],
    output_dir: Path,
    dry_run: bool = False,
    sheet_name: str = SHEET_NAME,
) -> dict[str, Any]:
    scan = scan_destination(dst_path, sheet_name)
    headers: HeaderMap = scan["headers"]
    rows_by_code: dict[str, int] = scan["rows_by_code"]
    duplicate_destination_codes = scan["duplicate_destination_codes"]

    unmatched_source = sorted(code for code in counts if code not in rows_by_code)
    destination_without_source = sorted(code for code in rows_by_code if code not in counts)
    applied_codes = sorted(code for code in counts if code in rows_by_code)

    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / dst_path.name
    backup_path = output_dir / "backup_original.xlsx"
    report_path = output_dir / "validation_report.json"

    report: dict[str, Any] = {
        "ok": False,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "dry_run": dry_run,
        "destination_filename": dst_path.name,
        "output_filename": output_path.name,
        "hashes": {
            "destination_original_sha256": file_sha256(dst_path),
        },
        "headers": headers.__dict__,
        "counts_total": len(counts),
        "applied_count": len(applied_codes),
        "applied_codes": applied_codes,
        "unmatched_source_codes": unmatched_source,
        "destination_without_source_codes": destination_without_source,
        "duplicate_destination_codes": duplicate_destination_codes,
        "updated_cell_type_errors": [],
        "non_quantity_value_diffs": [],
    }

    if duplicate_destination_codes:
        report["error"] = "codigos duplicados en destino"
        write_report(report_path, report)
        return report
    if unmatched_source:
        report["error"] = "conteos con codigo no encontrado en destino"
        write_report(report_path, report)
        return report

    if dry_run:
        report["ok"] = True
        write_report(report_path, report)
        return report

    shutil.copy2(dst_path, backup_path)
    before = workbook_value_snapshot(dst_path)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    shutil.copy2(dst_path, tmp_path)

    wb = openpyxl.load_workbook(tmp_path, data_only=False)
    ws = wb[sheet_name]
    allowed = set()
    for code in applied_codes:
        row_idx = rows_by_code[code]
        cell = ws.cell(row=row_idx, column=headers.qty_col)
        cell.value = float(counts[code])
        cell.number_format = "0.###"
        allowed.add((sheet_name, row_idx, headers.qty_col))
    wb.save(tmp_path)

    after = workbook_value_snapshot(tmp_path)
    report["non_quantity_value_diffs"] = diff_snapshots(before, after, allowed)

    wb_check = openpyxl.load_workbook(tmp_path, data_only=False)
    ws_check = wb_check[sheet_name]
    type_errors = []
    for code in applied_codes:
        row_idx = rows_by_code[code]
        value = ws_check.cell(row=row_idx, column=headers.qty_col).value
        if not isinstance(value, (int, float)):
            type_errors.append({"code": code, "row": row_idx, "value": value, "type": type(value).__name__})
    report["updated_cell_type_errors"] = type_errors

    if report["non_quantity_value_diffs"]:
        report["error"] = "se detectaron cambios fuera de la columna Cantidad"
    elif type_errors:
        report["error"] = "celdas actualizadas no quedaron numericas"
    else:
        shutil.move(str(tmp_path), output_path)
        report["ok"] = True
        report["hashes"]["output_sha256"] = file_sha256(output_path)
        report["hashes"]["backup_original_sha256"] = file_sha256(backup_path)

    if tmp_path.exists():
        tmp_path.unlink()
    write_report(report_path, report)
    return report


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2, sort_keys=True)
        f.write("\n")


def build_output_dir(base: Path) -> Path:
    stamp = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    return base / stamp


def parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument("--source", type=Path, help="XLSX fuente con conteo fisico")
    src.add_argument("--from-d1", nargs=2, metavar=("ALMACEN", "FECHA"), help="Usar sesion D1 como fuente")
    p.add_argument("--destination", type=Path, required=True, help="XLSX plantilla destino de Xetux")
    p.add_argument("--output-base", type=Path, default=Path("inventarios/outputs"))
    p.add_argument("--output-dir", type=Path, help="Directorio exacto de salida")
    p.add_argument("--sheet", default=SHEET_NAME)
    p.add_argument("--dry-run", action="store_true")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    if args.source:
        source_info = parse_source_xlsx(args.source, args.sheet)
    else:
        source_info = parse_d1_session(args.from_d1[0], args.from_d1[1])

    counts = source_info["counts"]
    if source_info.get("duplicate_source_codes"):
        report = {
            "ok": False,
            "error": "codigos duplicados en fuente",
            "duplicate_source_codes": source_info["duplicate_source_codes"],
        }
        out = args.output_dir or build_output_dir(args.output_base)
        write_report(out / "validation_report.json", report)
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 2
    if source_info.get("invalid_source_quantities"):
        report = {
            "ok": False,
            "error": "cantidades invalidas en fuente",
            "invalid_source_quantities": source_info["invalid_source_quantities"],
        }
        out = args.output_dir or build_output_dir(args.output_base)
        write_report(out / "validation_report.json", report)
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return 2

    output_dir = args.output_dir or build_output_dir(args.output_base)
    report = apply_counts(args.destination, counts, output_dir, args.dry_run, args.sheet)
    report["source"] = {k: v for k, v in source_info.items() if k != "counts"}
    write_report(output_dir / "validation_report.json", report)
    print(json.dumps({
        "ok": report["ok"],
        "output_dir": str(output_dir),
        "output_file": report.get("output_filename") if report["ok"] and not args.dry_run else None,
        "applied_count": report.get("applied_count"),
        "unmatched_source_codes": report.get("unmatched_source_codes"),
        "destination_without_source_count": len(report.get("destination_without_source_codes") or []),
        "error": report.get("error"),
    }, ensure_ascii=False, indent=2))
    return 0 if report["ok"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
