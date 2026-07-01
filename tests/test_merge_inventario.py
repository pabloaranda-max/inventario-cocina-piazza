import tempfile
import unittest
from pathlib import Path

import openpyxl

from tools.merge_inventario import apply_counts, parse_source_xlsx


def make_book(path: Path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(["#", "Tipo", "Grupo", "Subgrupo", "Código", "Artículo", "Unidad", "Cantidad"])
    for row in rows:
        ws.append(row)
    wb.save(path)


class MergeInventarioTest(unittest.TestCase):
    def test_parse_source_skips_empty_zero_and_stringifies_numeric(self):
        with tempfile.TemporaryDirectory() as tmp:
            src = Path(tmp) / "source.xlsx"
            make_book(src, [
                [1, "", "", "", "A1", "Uno", "PZA", "36"],
                [2, "", "", "", "A2", "Dos", "PZA", 0],
                [3, "", "", "", "A3", "Tres", "PZA", None],
            ])
            parsed = parse_source_xlsx(src)
            self.assertEqual(parsed["counts"], {"A1": 36.0})

    def test_apply_counts_only_changes_quantity_cells(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            dst = tmp / "Inventario_XTINV.xlsx"
            make_book(dst, [
                [1, "T", "G", "S", "A1", "Uno", "PZA", None],
                [2, "T", "G", "S", "A2", "Dos", "PZA", None],
                [3, "T", "G", "S", "XMAT_1234", "Pres", "PZA", None],
            ])
            report = apply_counts(dst, {"A1": 2.5}, tmp / "out")
            self.assertTrue(report["ok"], report)
            self.assertEqual(report["non_quantity_value_diffs"], [])
            out = tmp / "out" / dst.name
            wb = openpyxl.load_workbook(out, data_only=False)
            ws = wb["Inventario"]
            self.assertEqual(ws["H2"].value, 2.5)
            self.assertIsNone(ws["H3"].value)
            self.assertIsNone(ws["H4"].value)

    def test_unmatched_source_fails(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp = Path(tmp)
            dst = tmp / "Inventario_XTINV.xlsx"
            make_book(dst, [[1, "T", "G", "S", "A1", "Uno", "PZA", None]])
            report = apply_counts(dst, {"MISSING": 1}, tmp / "out")
            self.assertFalse(report["ok"])
            self.assertEqual(report["unmatched_source_codes"], ["MISSING"])


if __name__ == "__main__":
    unittest.main()
