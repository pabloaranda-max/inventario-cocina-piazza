#!/usr/bin/env python3
"""Tokens {{NOMBRE}} para la prosa de los entregables.

La prosa no lleva cifras escritas a mano: lleva tokens que se sustituyen al
construir, desde las cifras canónicas de datos.py. Si un número cambia, cambia
en todos lados o el build falla.
"""
import re
import datos


def construir(filas, T):
    tk = {}

    def add(nombre, valor, fmt='{:,.0f}'):
        tk[nombre] = fmt.format(valor)

    # consolidado
    add('NETO', T['neto']);            add('REAL_BRUTO', T['real_bruto'])
    add('TEORICO', T['teorico']);      add('BRECHA', T['brecha'])
    add('VENTA_ATRIB', T['venta']);    add('INICIAL', T['inicial'])
    add('COMPRAS', T['compras']);      add('FINAL', T['final'])
    add('INTRA', T['intra']);          add('DIF_VENTA', T['dif_venta'])
    add('VENTA_POS', datos.VENTA_POS); add('SIN_RECETA', datos.VENTA_SIN_RECETA)
    add('FALTANTE', T['costo_faltante'])
    add('SALUM_BRECHA', datos.SALUMERIA['brecha'])
    add('BRECHA_SALUM', T['brecha_con_salumeria'])
    add('COSTO_POS', datos.COSTO_POS)
    add('CAPITAL_DIA', T['capital_dia'])
    add('TRANSF_TOTAL', T['transf_total'])
    tk['MOVIMIENTOS'] = f'{T["movimientos"]}'

    # porcentajes
    add('PCT_REAL', T['pct_real'], '{:.1f}')
    add('PCT_TEO', T['pct_teo'], '{:.1f}')
    add('PTS', T['pts'], '{:.1f}')
    add('PCT_BRECHA_CONSUMO', T['pct_brecha_del_consumo'], '{:.1f}')
    add('PCT_MENU', datos.RATIO_COSTEADOS * 100, '{:.1f}')
    add('PCT_REGISTRADO', datos.COSTO_POS / datos.VENTA_POS * 100, '{:.1f}')
    add('PCT_SIN_RECETA', datos.VENTA_SIN_RECETA / datos.VENTA_POS * 100, '{:.1f}')
    add('PCT_TEO_EXPLICA', T['teorico'] / T['neto'] * 100, '{:.0f}')
    add('DIAS_INV', T['dias_inv'], '{:.0f}')
    add('ROT', T['rot'], '{:.1f}')

    # por almacén
    for f in filas:
        a = f['alm']
        add(f'{a}_NETO', f['neto']);     add(f'{a}_TEO', f['teorico'])
        add(f'{a}_BRECHA', abs(f['brecha']))
        add(f'{a}_FINAL', f['final'])
        add(f'{a}_DIAS', f['dias_inv'], '{:.0f}')
        add(f'{a}_ROT', f['rot'], '{:.1f}')
        if f['pct_real'] is not None:
            add(f'{a}_PCT', f['pct_real'], '{:.1f}')
        add(f'{a}_SHARE', abs(f['brecha']) / T['brecha'] * 100, '{:.0f}')

    # derivados de escenarios
    add('BRECHA_SIN_DESECHABLES', T['brecha'] - 12994.45)
    return tk


_TOKEN = re.compile(r'\{\{([A-Z_]+)\}\}')


def render(texto, tk, origen=''):
    """Sustituye {{TOKEN}}; falla si alguno no existe o queda sin sustituir."""
    faltantes = sorted({m for m in _TOKEN.findall(texto) if m not in tk})
    if faltantes:
        raise KeyError(f'{origen}: tokens sin definir -> {faltantes}')
    out = _TOKEN.sub(lambda m: tk[m.group(1)], texto)
    assert '{{' not in out, f'{origen}: quedaron tokens sin sustituir'
    return out


# cifras que ya no deben aparecer en ningún entregable
OBSOLETAS = ['647,862', '456,426', '191,436', '1,777,746', '566,417', '323,999',
             '267,819', '210,024', '57,794', '36.4 %', '25.7 %', '10.8 pts',
             '20,899', '262 días', '1,051,086', '780,425', '627,810', '767,249']


def _texto_de(ruta):
    """Contenido legible de un entregable, sea texto plano o libro de Excel."""
    if ruta.endswith('.xlsx'):
        from openpyxl import load_workbook
        wb = load_workbook(ruta)
        partes = []
        for hoja in wb.worksheets:
            for fila in hoja.iter_rows():
                for c in fila:
                    if c.value is not None:
                        partes.append(str(c.value))
        return ' '.join(partes)
    return open(ruta, encoding='utf-8').read()


def auditar(rutas):
    """Devuelve {archivo: [cifras obsoletas encontradas]}."""
    malos = {}
    for r in rutas:
        s = _texto_de(r)
        hits = [v for v in OBSOLETAS if v in s]
        if hits:
            malos[r] = hits
    return malos
