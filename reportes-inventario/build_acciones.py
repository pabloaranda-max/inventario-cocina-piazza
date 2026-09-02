#!/usr/bin/env python3
"""Listas de trabajo: una hoja por acción, con columnas para asignar y marcar."""
import xlrd, collections, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import datos

D = '/mnt/c/Users/P_ara/Downloads/'
OUT = '/home/lilp/proyectos/acciones-inventario-julio-2026.xlsx'
GREEN, WINE, BONE, PAPER, INK = '2F3E1E', '9B1E21', 'EEE3CA', 'FFFDF8', '1F2916'
MONEY = '#,##0.00;[Red]-#,##0.00'


def M(s):
    s = str(s).replace('$', '').replace(',', '').strip()
    n = s.startswith('-'); s = s.lstrip('-')
    try: v = float(s)
    except ValueError: v = 0.0
    return -v if n else v


def N(s):
    try: return float(str(s).replace(',', '').strip())
    except ValueError: return 0.0


def rows_of(f):
    sh = xlrd.open_workbook(D + f).sheets()[0]
    return [[str(c.value) for c in sh.row(r)] for r in range(sh.nrows)]


def variaciones(f):
    R = []
    for v in rows_of(f)[1:]:
        R.append(dict(cod=v[4].strip(), des=v[5].strip(), gr=v[1].strip(), um=v[6],
                      ini=N(v[7]), cini=M(v[8]), ent=N(v[9]), cent=M(v[10]),
                      fin=N(v[13]), cfin=M(v[14]), var=N(v[19]), cvar=M(v[21])))
    return R


wb = Workbook()
thin = Side(style='thin', color='D8CFB8')


def hoja(titulo, encabezado, filas, anchos, nota, numfmt=None, con_gestion=True):
    ws = wb.create_sheet(titulo)
    ws.cell(1, 1, nota).font = Font(name='Calibri', size=9, italic=True, color='6B6250')
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(encabezado) + (3 if con_gestion else 0))
    ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 30

    cols = list(encabezado) + (['Responsable', 'Fecha', 'Estado'] if con_gestion else [])
    for c, h in enumerate(cols, 1):
        cel = ws.cell(2, c, h)
        cel.font = Font(name='Calibri', bold=True, size=10, color=PAPER)
        cel.fill = PatternFill('solid', fgColor=GREEN if c <= len(encabezado) else WINE)
        cel.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = ws.cell(3, 1)

    for i, fila in enumerate(filas):
        for c, val in enumerate(fila, 1):
            cel = ws.cell(3 + i, c, val)
            cel.border = Border(bottom=thin)
            if numfmt and c in numfmt:
                cel.number_format = numfmt[c]
    if con_gestion and filas:
        dv = DataValidation(type='list', formula1='"pendiente,en proceso,hecho,no aplica"',
                            allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(len(encabezado) + 3)
        dv.add(f'{col}3:{col}{2 + len(filas)}')
        for i in range(len(filas)):
            ws.cell(3 + i, len(encabezado) + 3, 'pendiente')

    for c, w in enumerate(anchos + ([16, 12, 13] if con_gestion else []), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}{2 + len(filas)}'
    ws.sheet_view.showGridLines = False
    return ws


# ---------------------------------------------------------------- 1. sin receta
PROD = [dict(cod=v[1].strip(), nom=v[2].strip(), cant=N(v[3]), venta=M(v[5]), cprom=M(v[13]))
        for v in rows_of(datos.XLS_PRODUCTOS)[1:]]
SIN = sorted([p for p in PROD if p['cprom'] == 0 and p['venta'] > 0], key=lambda x: -x['venta'])
acc = 0.0
filas = []
for i, p in enumerate(SIN, 1):
    acc += p['venta']
    filas.append([i, p['cod'], p['nom'], p['cant'], p['venta'],
                  p['venta'] * datos.RATIO_COSTEADOS, acc / sum(q['venta'] for q in SIN) * 100])
hoja('1 · Sin receta', ['#', 'Código', 'Producto', 'Piezas', 'Venta neta',
                        'Costo estimado', '% acumulado'], filas,
     [5, 18, 46, 10, 14, 14, 12],
     f'ACCIÓN DE MAYOR RETORNO. {len(SIN)} productos vendieron ${sum(p["venta"] for p in SIN):,.0f} sin descontar inventario. '
     f'Cargar receta a cada uno. Están ordenados por venta: los 10 primeros son el 68 % del total, empieza por arriba. '
     f'El costo estimado aplica {datos.RATIO_COSTEADOS*100:.2f} %, la razón real de los productos costeados.',
     numfmt={5: MONEY, 6: MONEY, 7: '0.0"%"'})

# ---------------------------------------------------------------- 2. COCINA
COC = variaciones(datos.XLS_VARIACIONES_COCINA)
falt = sorted([x for x in COC if x['cvar'] < 0], key=lambda z: z['cvar'])
filas = [[i, x['cod'], x['des'], x['gr'], x['um'], abs(x['var']), abs(x['cvar'])]
         for i, x in enumerate(falt, 1)]
hoja('2 · COCINA faltantes', ['#', 'Código', 'Artículo', 'Grupo', 'UM',
                              'Cantidad faltante', 'Costo faltante'], filas,
     [5, 18, 44, 22, 8, 14, 14],
     f'COCINA concentra el 44 % de la brecha del mes (${[f for f in datos.calcular()[0] if f["alm"]=="COCINA"][0]["brecha"]:,.0f}). '
     f'{len(falt)} artículos con faltante, ${abs(sum(x["cvar"] for x in falt)):,.0f} en total. Los primeros son lácteos y carnes que se '
     f'porcionan: revisar si la receta descuenta lo que realmente se usa. Ojo: COCINA solo contó 69 códigos de 742, '
     f'así que parte de esto es conteo que no se hizo, no merma.',
     numfmt={6: '#,##0.00', 7: MONEY})

# ---------------------------------------------------------------- 3. GENERAL sin contar
GEN = variaciones(datos.XLS_VARIACIONES_GENERAL)
nc = sorted([x for x in GEN if x['fin'] == 0 and (x['cini'] + x['cent']) > 0],
            key=lambda z: -(z['cini'] + z['cent']))
filas = [[i, x['cod'], x['des'], x['gr'], x['um'], x['ini'], x['cini'], x['cent']]
         for i, x in enumerate(nc, 1)]
hoja('3 · GENERAL sin contar', ['#', 'Código', 'Artículo', 'Grupo', 'UM',
                                'Existencia inicial', 'Costo inicial', 'Compras del mes'], filas,
     [5, 18, 44, 20, 8, 13, 13, 13],
     f'{len(nc)} códigos con existencia o compras cuyo conteo final quedó en CERO: el 79 % del residuo de GENERAL. '
     f'Casi todo es empaque y desechables. Decisión: sacarlos del inventario valorizado (gasto directo en la compra) '
     f'Y quitarlos de las recetas. A medias, su compra completa cae en la variación cada mes.',
     numfmt={7: MONEY, 8: MONEY})

# ---------------------------------------------------------------- 4. ALIMENTARI
ALI = variaciones(datos.XLS_VARIACIONES_ALIMENTARI)
alifalt = sorted([x for x in ALI if x['cvar'] < 0], key=lambda z: z['cvar'])
filas = [[i, x['cod'], x['des'], x['gr'], x['um'], x['fin'], x['cfin'], abs(x['cvar'])]
         for i, x in enumerate(alifalt, 1)]
_f = [f for f in datos.calcular()[0] if f['alm'] == 'ALIMENTARI'][0]
hoja('4 · ALIMENTARI rotación', ['#', 'Código', 'Artículo', 'Grupo', 'UM',
                                 'Existencia final', 'Costo final', 'Faltante'], filas,
     [5, 18, 44, 20, 8, 13, 13, 13],
     f'ALIMENTARI tiene {_f["dias_inv"]:.0f} días de inventario y {_f["rot"]:.1f} rotaciones al año: ${_f["final"]:,.0f} de capital parado. '
     f'No es un problema de brecha —consumió menos de lo esperado— sino de dinero inmovilizado. '
     f'Revisar qué de esto no rota y decidir: liquidar, promocionar o dejar de comprar.',
     numfmt={7: MONEY, 8: MONEY})

# ---------------------------------------------------------------- 5. checklist
filas_, T, _ = datos.calcular()
coc = [f for f in filas_ if f['alm'] == 'COCINA'][0]
chk = [
    [1, 'Cargar receta a los productos sin costeo', f'{len(SIN)} productos · hoja 1',
     sum(p['venta'] for p in SIN) * datos.RATIO_COSTEADOS, 'Cocina + sistemas'],
    [2, 'Auditar COCINA artículo por artículo', f'{len(falt)} artículos · hoja 2', coc['brecha'], 'Cocina'],
    [3, 'Reaplicar SALUMERIA en Xetux', 'Verificar que COCINA dé $367,510 de costo real',
     datos.SALUMERIA['brecha'], 'Contraloría'],
    [4, 'Recalibrar las recetas que descuentan de menos', '135 artículos, medido sobre una semana', 51333, 'Cocina'],
    [5, 'Auditar CAVA', '12 artículos con receta activa sin venta asociada',
     [f for f in filas_ if f['alm'] == 'CAVA'][0]['brecha'], 'Cava + auditoría'],
    [6, 'Capturar merma y consumo interno', 'Hoy son 2 registros en todo el mes', 35000, 'Operación'],
    [7, 'Sacar desechables del inventario valorizado', f'{len(nc)} códigos · hoja 3', 12994, 'Contraloría'],
    [8, 'Revisar capital parado en ALIMENTARI', f'{_f["dias_inv"]:.0f} días de inventario · hoja 4',
     _f['final'], 'Alimentari'],
    [9, 'Corregir atribución de venta de salones Amici', 'Normaliza COCINA y AMICI a la vez', None, 'Sistemas'],
    [10, 'Configurar bien el semanal para que no se guarde como mensual',
     'Es lo que partió el periodo de COCINA y ALIMENTARI en julio', None, 'Contraloría'],
]
hoja('Acciones', ['#', 'Acción', 'Detalle', 'Monto en juego', 'Responsable sugerido'], chk,
     [5, 46, 52, 15, 22],
     f'Ordenadas por retorno. El monto es lo que está en juego en la partida, no un ahorro garantizado. '
     f'Brecha total del mes: ${T["brecha"]:,.0f}, que sube a ~${T["brecha_con_salumeria"]:,.0f} al reaplicar SALUMERIA.',
     numfmt={4: MONEY})

wb.move_sheet('Acciones', offset=-5)
del wb['Sheet']
wb.save(OUT)
print(f'{os.path.basename(OUT)}  {os.path.getsize(OUT)/1024:.0f} KB')
for n in wb.sheetnames:
    print(f'   {n}: {wb[n].max_row - 2} renglones')
