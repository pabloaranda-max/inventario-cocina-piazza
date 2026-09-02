#!/usr/bin/env python3
"""Construye el Excel de trabajo del costo de inventario de julio 2026."""
import xlrd, collections, re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

D = '/mnt/c/Users/P_ara/Downloads/'
OUT = '/home/lilp/proyectos/costo-inventario-julio-2026.xlsx'

GREEN, WINE, BONE, YELLOW, INK, PAPER = '2F3E1E', '9B1E21', 'EEE3CA', 'EFA91E', '1F2916', 'FFFDF8'

def M(s):
    s = str(s).replace('$', '').replace(',', '').strip()
    neg = s.startswith('-'); s = s.lstrip('-')
    try: v = float(s)
    except ValueError: v = 0.0
    return -v if neg else v

def N(s):
    try: return float(str(s).replace(',', '').strip())
    except ValueError: return 0.0

def rows_of(fname, sheet=0):
    sh = xlrd.open_workbook(D + fname).sheets()[sheet]
    return [[str(c.value) for c in sh.row(r)] for r in range(sh.nrows)]

# ---------------------------------------------------------------- datos fuente
ALM = {  # inicial, compras/transf+, final, desperdicio, cons.interno, real, teorico, ventas/transf-
    'GENERAL':    (208869.55, 141906.81, 177180.17,   0.00, 62.00, 173534.19,   6405.71,  142155.20),
    'COCINA':     ( 94131.66, 344747.48,  71369.04,   0.00,  0.00, 367510.10, 259459.18, 1278621.24),
    'BARRA':      (131712.27, 122980.61, 117423.79,   0.00,  0.00, 137269.09, 117605.54,  345647.46),
    'AMICI':      ( 90584.95,  81152.25,  95361.93,   0.00,  0.00,  76375.27,  24849.10,   81237.72),
    'CAVA':       (178936.49,  34079.75,  93466.75, 112.50,  0.00, 119436.99,  84995.12,  213241.64),
    'ALIMENTARI': ( 62819.02,   8704.31,  59657.61,   0.00,  0.00,  11865.72,  16271.28,   92501.53),
}
SALUMERIA = (86786.51, 52311.75, 76847.32, 0.00, 0.00, 62250.94, 8188.37, 0.00)
INICIO = {'GENERAL': '30/06/2026', 'COCINA': '30/06/2026 21:39', 'BARRA': '30/06/2026 23:59',
          'AMICI': '30/06/2026 21:45', 'CAVA': '30/06/2026 23:59', 'ALIMENTARI': '30/06/2026 21:48'}
PERIM = set(ALM)

# transferencias detalladas
TR = []
for v in rows_of('Reporte de Transferencia entre Almacenes-Sat Aug 01 05_42_08 CST 2026.xls')[1:]:
    TR.append(dict(orig=v[1].strip(), dest=v[2].strip(), grupo=v[3].strip(), folio=v[4],
                   art=v[7].strip(), cant=N(v[8]), um=v[9], costo=M(v[11]),
                   encargado=v[12].strip(), fecha=v[13].strip()))
intra = collections.defaultdict(float)
for t in TR:
    if t['orig'] in PERIM and t['dest'] in PERIM:
        intra[t['orig']] += t['costo']
TI = sum(intra.values())

# variaciones GENERAL
GEN = []
for v in rows_of('Variaciones de Inventario - XTINV000285.xls')[1:]:
    GEN.append(dict(grupo=v[1].strip(), subgrupo=v[2].strip(), cod=v[4].strip(), des=v[5].strip(),
                    um=v[6], ini=N(v[7]), cini=M(v[8]), ent=N(v[9]), cent=M(v[10]),
                    sal=N(v[11]), csal=M(v[12]), fin=N(v[13]), cfin=M(v[14]), cur=M(v[16])))

# productos vendidos
PROD = []
for v in rows_of('Detallado por Producto-Sat Aug 01 04_59_00 CST 2026.xls')[1:]:
    PROD.append(dict(cod=v[1].strip(), nom=v[2].strip(), cant=N(v[3]), venta=M(v[5]),
                     desc=M(v[7]), ult=M(v[9]), cprom=M(v[13])))
SIN = sorted([p for p in PROD if p['cprom'] == 0 and p['venta'] > 0], key=lambda x: -x['venta'])
CON = [p for p in PROD if p['cprom'] > 0]
RATIO = sum(p['cprom'] for p in CON) / sum(p['venta'] for p in CON)

# ambientes
AMB = collections.defaultdict(lambda: [0.0, 0.0])
for v in rows_of('Detallado por Producto-Sat Aug 01 05_07_25 CST 2026.xls')[1:]:
    AMB[v[2].strip()][0] += M(v[8]); AMB[v[2].strip()][1] += M(v[16])

VENTA = sum(p['venta'] for p in PROD)
COSTO_POS = sum(p['cprom'] for p in PROD)
REAL = sum(ALM[k][5] for k in ALM)
TEO = sum(ALM[k][6] for k in ALM)

# ---------------------------------------------------------------- estilos
wb = Workbook()
thin = Side(style='thin', color='D8CFB8')
H = dict(font=Font(name='Calibri', bold=True, size=10, color=PAPER),
         fill=PatternFill('solid', fgColor=GREEN),
         align=Alignment(horizontal='center', vertical='center', wrap_text=True))
MONEY = '#,##0.00;[Red]-#,##0.00'
PCT = '0.0"%"'

def sheet(title, headers, data, widths, note=None, numfmt=None, total_row=None):
    ws = wb.create_sheet(title)
    r = 1
    if note:
        ws.cell(1, 1, note).font = Font(name='Calibri', size=9, italic=True, color='6B6250')
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.row_dimensions[1].height = 28
        ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical='center')
        r = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font, cell.fill, cell.alignment = H['font'], H['fill'], H['align']
    ws.row_dimensions[r].height = 30
    ws.freeze_panes = ws.cell(r + 1, 1)
    for i, row in enumerate(data):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r + 1 + i, c, val)
            cell.border = Border(bottom=thin)
            if numfmt and c in numfmt:
                cell.number_format = numfmt[c]
            if isinstance(val, str) and c == 1:
                cell.font = Font(name='Calibri', size=10)
    if total_row:
        rr = r + 1 + len(data)
        for c, val in enumerate(total_row, 1):
            cell = ws.cell(rr, c, val)
            cell.font = Font(name='Calibri', bold=True, size=10, color=GREEN)
            cell.border = Border(top=Side(style='medium', color=GREEN))
            if numfmt and c in numfmt:
                cell.number_format = numfmt[c]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.sheet_view.showGridLines = False
    return ws

# ---------------------------------------------------------------- 1. RESUMEN
ws = wb.active; ws.title = 'Resumen'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 46
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 58

def block(r, title):
    c = ws.cell(r, 2, title)
    c.font = Font(name='Calibri', bold=True, size=11, color=PAPER)
    c.fill = PatternFill('solid', fgColor=GREEN)
    for col in (3, 4, 5):
        ws.cell(r, col).fill = PatternFill('solid', fgColor=GREEN)
    ws.row_dimensions[r].height = 22
    return r + 1

def line(r, label, val, fmt=MONEY, note='', bold=False, color=None):
    ws.cell(r, 2, label).font = Font(name='Calibri', size=10, bold=bold,
                                     color=color or INK)
    c = ws.cell(r, 3, val)
    c.number_format = fmt
    c.font = Font(name='Calibri', size=10, bold=bold, color=color or INK)
    c.alignment = Alignment(horizontal='right')
    if note:
        n = ws.cell(r, 5, note)
        n.font = Font(name='Calibri', size=9, color='6B6250')
        n.alignment = Alignment(wrap_text=True, vertical='center')
    return r + 1

ws.cell(2, 2, 'COSTO DE INVENTARIO — JULIO 2026').font = Font(name='Calibri', bold=True, size=18, color=GREEN)
ws.cell(3, 2, 'Piazza Pasticcio · 6 de 7 almacenes aplicados · SALUMERIA pendiente de reaplicación').font = Font(name='Calibri', size=10, color='6B6250')

VENTA_SIN = sum(p['venta'] for p in SIN)
FALTANTE = VENTA_SIN * RATIO
VENTA_ATRIB = sum(ALM[k][7] for k in ALM) - sum(t['costo'] for t in TR if t['orig'] in PERIM)
REAL_NETO = REAL - TI

r = 5
r = block(r, 'EL COSTO REAL DEL MES')
r = line(r, 'COSTO REAL DE CONSUMO', REAL_NETO / VENTA_ATRIB * 100, fmt=PCT, bold=True, color=WINE,
         note='Lo que SALIÓ del inventario: inicial + compras − contado físicamente al cierre. No depende de las recetas.')
r = line(r, 'Costo teórico, por receta', TEO / VENTA_ATRIB * 100, fmt=PCT,
         note='Lo que las recetas dicen que DEBIÓ salir por lo que se vendió.')
r = line(r, 'BRECHA, en puntos', (REAL_NETO - TEO) / VENTA_ATRIB * 100, fmt=PCT, bold=True, color=WINE)
r += 1
r = line(r, '  · costo real consolidado', REAL_NETO)
r = line(r, '  · costo teórico', TEO)
r = line(r, '  · BRECHA en pesos', REAL_NETO - TEO, bold=True, color=WINE)
r = line(r, '  · venta atribuida (denominador)', VENTA_ATRIB,
         note='NO es la venta del mes. El mes vendió $2,344,163. Esta cifra es la venta que Xetux asigna a los seis '
              'almacenes, que es la venta de los productos con receta —los únicos que consumen de un almacén—. Es el denominador '
              'correcto. La diferencia de $324,172 son los $312,945 de productos sin receta, que no descuentan de ningún inventario '
              'y por eso no se asignan a ningún almacén.')

r += 1
r = block(r, '¿Y EL MENÚ ESTÁ BIEN CONSTRUIDO? — sí, es otra pregunta')
r = line(r, 'Costo del menú por receta (POS)', RATIO * 100, fmt=PCT, bold=True, color=GREEN,
         note='Sobre los 409 productos costeados. Coincide con el 25.2 % teórico de los almacenes: '
              'dos mediciones independientes que se validan entre sí.')
r = line(r, 'Venta neta del mes (POS)', VENTA, note='Mes completo, 01–31/07.')
r = line(r, 'Costo por receta registrado', COSTO_POS)
r = line(r, 'Venta sin costo asociado', VENTA_SIN, color=WINE,
         note=f'{VENTA_SIN/VENTA*100:.1f} % de la venta, en {len(SIN)} productos. Ver hoja "Botones sin receta".')
r = line(r, 'Costo faltante estimado', FALTANTE, color=WINE)
r = line(r, 'NO USAR: costo registrado / venta total', COSTO_POS / VENTA * 100, fmt=PCT, color=WINE,
         note='Subestimado. Sale bajo porque 73 productos venden con costo cero. Ni este ni el 25.4 % '
              'son el costo real: los dos son teóricos.')

r += 1
r = block(r, 'LA BRECHA, CORREGIDA')
r = line(r, 'Costo real — suma de los 6 almacenes', REAL, note='Como lo reporta Xetux almacén por almacén.')
r = line(r, 'menos transferencias internas', -TI, color=WINE,
         note='Doble conteo: salen de un almacén y entran a otro, el sistema las cuenta dos veces.')
r = line(r, 'COSTO REAL CONSOLIDADO', REAL - TI, bold=True)
r = line(r, 'Costo teórico', TEO)
r = line(r, 'BRECHA REAL', REAL - TI - TEO, bold=True, color=WINE,
         note='Las transferencias entre almacenes contados aparecen como salida del emisor y entrada del receptor.')
r = line(r, 'Cobertura del teórico sobre el consumo real', TEO / (REAL - TI) * 100, fmt=PCT)

r += 1
r = block(r, 'ALCANCE Y PENDIENTES')
r = line(r, 'SALUMERIA — brecha si se reaplica con cifras actuales', SALUMERIA[5] - SALUMERIA[6],
         note='No está en ningún total. Su corte quedó anulado.')
r = line(r, 'Brecha proyectada con SALUMERIA', REAL - TI - TEO + SALUMERIA[5] - SALUMERIA[6])
r = line(r, 'Efecto de sacar desechables del inventario valorizado', -12994.45, color=GREEN,
         note='Ver hoja GENERAL. Requiere sacarlos del valorizado Y de las recetas.')

r += 2
ws.cell(r, 2, 'Los periodos no son iguales entre almacenes: cada uno corre desde su última toma aplicada. '
              'COCINA y ALIMENTARI arrancan el 08/07 (23.5 días), los otros cuatro el 30/06. Es lo normal con tomas '
              'escalonadas. El 31/07 fue la primera vez que los 7 cerraron con el mismo sello de tiempo, así que de '
              'agosto en adelante los periodos quedan alineados.').font = Font(name='Calibri', size=9, italic=True, color='6B6250')
ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=5)
ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')

# ---------------------------------------------------------------- 2. POR ALMACEN
data = []
orden = sorted(ALM, key=lambda x: -(ALM[x][5] - intra[x] - ALM[x][6]))
for i, k in enumerate(orden):
    R_ = 3 + i  # fila de Excel: encabezado en 2, datos desde 3
    den = ALM[k][7] - sum(t['costo'] for t in TR if t['orig'] == k)
    data.append([k, INICIO[k], ALM[k][0], ALM[k][1], ALM[k][2], ALM[k][5], intra[k],
                 f'=F{R_}-G{R_}', ALM[k][6], f'=H{R_}-I{R_}', den,
                 f'=IF(K{R_}>1000,H{R_}/K{R_}*100,"")'])
L, U = 3, 2 + len(orden)
tot = ['TOTAL', ''] + [f'=SUM({c}{L}:{c}{U})' for c in 'CDEFG'] + \
      [f'=F{U+1}-G{U+1}', f'=SUM(I{L}:I{U})', f'=H{U+1}-I{U+1}', f'=SUM(K{L}:K{U})',
       f'=H{U+1}/K{U+1}*100']
sheet('Brecha por almacén',
      ['Almacén', 'Periodo desde', 'Inv. inicial', 'Compras y Transf(+)', 'Inv. final contado',
       'Costo real Xetux', 'Transf. internas (doble conteo)', 'Costo real neto', 'Costo teórico',
       'BRECHA CORREGIDA', 'Venta atribuida', '% sobre venta'],
      data, [16, 17, 14, 16, 15, 15, 16, 14, 14, 16, 15, 12],
      note='"Venta atribuida" NO es la venta del mes: el mes vendió $2,344,163 y esta columna suma $2,019,991, porque es '
           'las mismas ventanas de tiempo que el costo de cada almacén. Es el denominador correcto de los porcentajes. '
           'Las columnas Costo real neto, BRECHA y % son fórmulas: si corriges un dato, se recalculan. '
           'GENERAL no vende: su % no es food cost y por eso queda en blanco. AMICI al 94 % es artefacto de denominador '
           '— su almacén recibe $81,238 de los $217,272 vendidos en sus ambientes; normalizado corre a ~35 %.',
      numfmt={3: MONEY, 4: MONEY, 5: MONEY, 6: MONEY, 7: MONEY, 8: MONEY, 9: MONEY, 10: MONEY, 11: MONEY, 12: PCT},
      total_row=tot)

# ---------------------------------------------------------------- 3. GENERAL
g = collections.defaultdict(float); nsin = collections.Counter(); csin = collections.defaultdict(float)
ncod = collections.Counter()
for x in GEN:
    g[x['grupo']] += x['cur']; ncod[x['grupo']] += 1
    if x['fin'] == 0 and (x['cini'] + x['cent']) > 0:
        nsin[x['grupo']] += 1; csin[x['grupo']] += x['cur']
resid = sum(x['cur'] for x in GEN)
gord = sorted(g.items(), key=lambda kv: -kv[1])
GL, GU = 3, 2 + len(gord)
data = [[k, ncod[k], nsin[k], v, f'=D{3+i}/$D${GU+1}*100', csin[k]] for i, (k, v) in enumerate(gord)]
sheet('GENERAL por grupo',
      ['Grupo', 'Códigos', 'Códigos sin conteo final', 'Uso real sin explicar', '% del residuo',
       'del cual, de códigos no contados'],
      data, [26, 10, 16, 18, 13, 20],
      note='Residuo de GENERAL: consumo que no queda explicado por venta ni por transferencia. EMPAQUE (desechables) es el '
           '40 %: 80 de sus 86 códigos no recibieron conteo final. En total, $25,535 —el 79 % del residuo— sale de 59 códigos '
           'que tuvieron movimiento y se contaron como CERO. No es merma: es cobertura de conteo.',
      numfmt={4: MONEY, 5: PCT, 6: MONEY},
      total_row=['TOTAL', len(GEN), sum(nsin.values()), f'=SUM(D{GL}:D{GU})', 100.0,
                 f'=SUM(F{GL}:F{GU})'])

nc = sorted([x for x in GEN if x['fin'] == 0 and (x['cini'] + x['cent']) > 0], key=lambda z: -z['cur'])
sheet('GENERAL códigos sin contar',
      ['Código', 'Descripción', 'Grupo', 'UM', 'Inicial', 'Costo inicial', 'Entradas', 'Costo entradas',
       'Salidas', 'Costo salidas', 'Final contado', 'Cae en el residuo'],
      [[x['cod'], x['des'], x['grupo'], x['um'], x['ini'], x['cini'], x['ent'], x['cent'],
        x['sal'], x['csal'], x['fin'], x['cur']] for x in nc],
      [18, 42, 20, 8, 10, 13, 10, 13, 10, 13, 12, 15],
      note='59 códigos con existencia inicial o entradas del mes cuyo conteo final quedó en CERO. Esta es la lista que '
           'explica el 79 % del residuo de GENERAL.',
      numfmt={6: MONEY, 8: MONEY, 10: MONEY, 12: MONEY},
      total_row=['', f'{len(nc)} códigos', '', '', None, sum(x['cini'] for x in nc), None,
                 sum(x['cent'] for x in nc), None, sum(x['csal'] for x in nc), None, sum(x['cur'] for x in nc)])

# ---------------------------------------------------------------- 4. BOTONES SIN RECETA
sheet('Botones sin receta',
      ['#', 'Código', 'Producto', 'Piezas vendidas', 'Venta neta', '% de la venta del mes',
       'Costo estimado', 'Acumulado de venta'],
      [[i + 1, p['cod'], p['nom'], p['cant'], p['venta'], f'=E{3+i}/{VENTA:.2f}*100',
        f'=E{3+i}*{RATIO:.6f}', f'=SUM($E$3:E{3+i})'] for i, p in enumerate(SIN)],
      [5, 18, 44, 13, 14, 15, 14, 16],
      note=f'{len(SIN)} productos vendieron ${sum(p["venta"] for p in SIN):,.2f} con costo asociado $0.00 — se sirvieron sin '
           f'descontar un gramo de inventario. El costo estimado aplica {RATIO*100:.2f} %, que es lo que realmente corren los '
           f'productos sí costeados. Los 10 primeros concentran ${sum(p["venta"] for p in SIN[:10]):,.2f}. '
           f'Esta es la acción de mayor retorno: no requiere investigación, es una lista.',
      numfmt={5: MONEY, 6: PCT, 7: MONEY, 8: MONEY},
      total_row=[None, '', f'{len(SIN)} productos', f'=SUM(D3:D{2+len(SIN)})', f'=SUM(E3:E{2+len(SIN)})',
                 f'=SUM(F3:F{2+len(SIN)})', f'=SUM(G3:G{2+len(SIN)})', None])

# ---------------------------------------------------------------- 5. TRANSFERENCIAS
orig = sorted({t['orig'] for t in TR}); dest = sorted({t['dest'] for t in TR})
mat = collections.defaultdict(float)
for t in TR:
    mat[(t['orig'], t['dest'])] += t['costo']
sheet('Transferencias matriz',
      ['Origen →  Destino'] + dest + ['TOTAL SALE', 'Dentro del perímetro'],
      [[o] + [mat.get((o, d)) or None for d in dest] + [sum(t['costo'] for t in TR if t['orig'] == o), intra[o] or None]
       for o in orig],
      [26] + [13] * len(dest) + [14, 17],
      note='599 movimientos, $150,744.18, todos con origen, destino, encargado y fecha. GENERAL origina el 87 % y no recibe '
           'nada; COCINA absorbe el 61 %. La columna final es la parte que se contaba dos veces en el consolidado.',
      numfmt={i: MONEY for i in range(2, len(dest) + 4)},
      total_row=['TOTAL ENTRA'] + [sum(t['costo'] for t in TR if t['dest'] == d) for d in dest]
                + [sum(t['costo'] for t in TR), TI])

sheet('Transferencias detalle',
      ['Folio', 'Origen', 'Destino', 'Grupo', 'Artículo', 'Cantidad', 'UM', 'Costo total', 'Encargado', 'Fecha'],
      [[t['folio'], t['orig'], t['dest'], t['grupo'], t['art'], t['cant'], t['um'], t['costo'],
        t['encargado'], t['fecha']] for t in TR],
      [16, 24, 16, 18, 42, 10, 8, 13, 18, 15],
      note='Trazabilidad completa del mes. Una sola persona registró el 93 % de los movimientos (558 de 599): '
           'concentración operativa que conviene repartir.',
      numfmt={8: MONEY},
      total_row=[None, '', '', '', f'{len(TR)} movimientos', None, '', sum(t['costo'] for t in TR), '', ''])

# ---------------------------------------------------------------- 6. AMBIENTES
aord = sorted(AMB.items(), key=lambda kv: -kv[1][0])
AL, AU = 3, 2 + len(aord)
data = [[a, v, c, f'=C{3+i}/B{3+i}*100'] for i, (a, (v, c)) in enumerate(aord)]
sheet('Costo por ambiente',
      ['Ambiente', 'Venta neta', 'Costo por receta', '% de costo'],
      data, [26, 16, 16, 12],
      note='EVENTOS al 11.2 % es la mitad del salón restaurante y consume de cava, ambas barras y cocina. Sus dos botones '
           'genéricos —OTROS EVENTO y ALIMENTOS EVENTO— son $74,138 sin costear: ahí está la mayor parte de por qué se ve '
           'tan barato.',
      numfmt={2: MONEY, 3: MONEY, 4: PCT},
      total_row=['TOTAL', f'=SUM(B{AL}:B{AU})', f'=SUM(C{AL}:C{AU})', f'=C{AU+1}/B{AU+1}*100'])

# ---------------------------------------------------------------- 7. ACCIONES
acc = [
    ['1', 'Cargar receta a los 73 botones sin costeo', 'Los 10 primeros concentran $213,682 de venta sin costear. '
     'No requiere investigación previa: es una lista, está en la hoja "Botones sin receta".', 79450, 'Alta', 'Cocina + sistemas'],
    ['2', 'Sacar desechables y suministros del inventario valorizado', 'Decisión ya tomada: de agosto no se cuentan. Para que '
     'funcione hay que sacarlos del valorizado (gasto directo en la compra) Y quitarlos de las recetas. A medias, su compra '
     'completa cae en la variación todos los meses.', 12994, 'Alta', 'Contraloría'],
    ['3', 'Reaplicar SALUMERIA', 'Su corte quedó anulado. Con cifras actuales aporta $54,063 a la brecha.', 54063, 'Alta', 'Contraloría'],
    ['4', 'Recalibrar las 135 recetas que descuentan de menos', 'Harina 75 kg usados / 20 descontados · aceite de oliva 18.9 / 4.7 · '
     'pollo orgánico 23.0 / 8.5. Medido sobre el análisis semanal del 28/07; en el mes será mayor.', 51333, 'Alta', 'Cocina'],
    ['5', 'Auditar CAVA', 'Único almacén sin explicación estructural: no transforma producto ni prepara mezclas. '
     '12 artículos con receta activa salieron sin venta asociada, $17,994.', 34442, 'Media', 'Cava + auditoría'],
    ['6', 'Capturar merma y consumo interno', 'Todo el mes tiene 2 registros por $174.50. Lo esperable a 1.5 % de venta es ~$35,000. '
     'Existe un ALMACEN COMIDA PERSONAL que mueve producto por transferencia en vez de registrarlo como consumo interno.', 35000, 'Media', 'Operación'],
    ['7', 'Corregir la atribución de venta de los salones Amici', 'Su comida se costea a COCINA. Ajustado, COCINA baja de 25.5 % a ~23.8 % '
     'y AMICI deja de verse al 94 %.', None, 'Media', 'Sistemas'],
    ['8', 'Subir la cobertura de conteo en GENERAL', '59 códigos con movimiento se contaron como cero. Cobertura de la toma: 165 de 434 códigos.', 25535, 'Media', 'Almacén'],
    ['9', 'Repartir el registro de transferencias', 'Una sola persona registró 558 de 599 movimientos del mes.', None, 'Baja', 'Operación'],
    ['10', 'Incorporar los 4 almacenes fuera de la toma', 'ALMACEN DE LIMPIEZA envía $13,159 · SALON recibe $5,223 · '
     'SUMINISTROS envía $3,453 · COMIDA PERSONAL envía $719. Ninguno tiene conteo.', 22553, 'Baja', 'Contraloría'],
    ['11', 'Preguntar a Xetux la inconsistencia entre sus dos reportes', 'Para GENERAL, el resumen por correo dice salidas $142,155 y el '
     'reporte de variaciones dice $147,002. Difieren $4,847 y mueven el residuo en $3,192.', None, 'Baja', 'Sistemas'],
]
sheet('Acciones',
      ['#', 'Acción', 'Por qué / detalle', 'Monto en juego', 'Prioridad', 'Responsable'],
      acc, [5, 46, 74, 15, 11, 20],
      note='Ordenadas por retorno. El monto es lo que está en juego en la partida, no un ahorro garantizado.',
      numfmt={4: MONEY})
for row in wb['Acciones'].iter_rows(min_row=3, min_col=3, max_col=3):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical='top')
        wb['Acciones'].row_dimensions[c.row].height = 46

# ---------------------------------------------------------------- 8. FUENTES
sheet('Fuentes y método',
      ['Concepto', 'Detalle'],
      [['Periodo', 'Venta 01–31/07/2026. Los almacenes corren desde su última toma aplicada (ver hoja Brecha por almacén).'],
       ['Fuente de los almacenes', 'Correos "Informe de Inventarios PIAZZA PASTICCIO" de xetux-noreply@xetux-e.com del 01/08/2026, uno por almacén.'],
       ['Fuente de la venta', 'Detallado por Producto de Xetux, 487 productos y 9 ambientes, mes completo.'],
       ['Fuente de transferencias', 'Reporte de Transferencia entre Almacenes con columnas de origen y destino, 599 movimientos. '
        'Ojo: existe otro export del mismo nombre que solo trae artículo y costo, sin origen ni destino — ese no sirve para reconciliar.'],
       ['Fuente del desglose de GENERAL', 'Variaciones de Inventario XTINV000285, 434 códigos.'],
       ['Corrección aplicada', 'El consolidado NO es la suma de los correos: las transferencias internas se cuentan dos veces '
        '(salida del emisor vía inventario final más bajo, entrada del receptor vía Compras/Transf(+)). Se netean $132,563.25.'],
       ['Cifra que se corrigió', 'GENERAL se reportó con $167,128 de brecha y $31,379 "sin rastro". Lo correcto es $36,417 de brecha. '
        'El $31,379 restaba la columna Ventas/Transf(-), que mezcla transferencias a costo con venta a precio de menú.'],
       ['Validación independiente', 'SALUMERIA: compra propia $52,023 cuadra al peso contra su reporte de proveedor (PRODUCTORES BARGE, $52,023.00).'],
       ['Estimaciones', 'Solo el costo de los botones sin receta está estimado, aplicando 25.39 %. Todo lo demás viene directo de Xetux.'],
       ['Denominador — cuidado', '"Venta atribuida" NO es la venta del mes. El mes vendió $2,344,162.90; la venta atribuida a los seis '
        'almacenes suma $2,019,991.18: es la venta de los productos que sí tienen receta, los únicos que consumen de un almacén, y por eso '
        'es el denominador correcto de los porcentajes. Difieren $324,172, que son los $312,945 de productos sin receta.'],
       ['Costo del menú — cuidado', 'El 22.0 % que resulta de dividir el costo registrado entre la venta total está SUBESTIMADO: '
        '73 productos ($312,945, el 13.3 % de la venta) tienen costo cero. El número defendible es 25.39 %, medido sobre los 409 '
        'productos que sí tienen receta cargada. No usar el 22 % hacia afuera.'],
       ['Fórmulas', 'Las columnas derivadas y los totales son fórmulas vivas: si corriges un dato de origen, se recalculan. '
        'Los datos de origen sí son valores fijos copiados de los reportes de Xetux.'],
       ['Qué NO es', 'No es una conciliación contable definitiva. Los periodos de COCINA y ALIMENTARI no cubren el mes completo, '
        'así que el consolidado mezcla ventanas de medición.'],
       ['Toma física', 'Conteo digital del 31/07/2026 en la app de inventarios, 16 personas, 7 almacenes, 24 rebanadas, 28 acuses inmutables.']],
      [30, 118],
      note='Todo el reporte es reproducible desde estos archivos.')
for row in wb['Fuentes y método'].iter_rows(min_row=3, min_col=2, max_col=2):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical='top')
        wb['Fuentes y método'].row_dimensions[c.row].height = 40

wb.save(OUT)
print('escrito:', OUT, os.path.getsize(OUT), 'bytes')
print('hojas:', wb.sheetnames)
print(f'venta {VENTA:,.2f} | costo POS {COSTO_POS:,.2f} ({COSTO_POS/VENTA*100:.2f}%)')
print(f'real {REAL:,.2f} - intra {TI:,.2f} = {REAL-TI:,.2f} | teorico {TEO:,.2f} | brecha {REAL-TI-TEO:,.2f}')
print(f'residuo GENERAL {resid:,.2f} | sin contar {sum(csin.values()):,.2f} | botones {len(SIN)} = {sum(p["venta"] for p in SIN):,.2f}')
